from fastapi import FastAPI, APIRouter, HTTPException
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
import os
import logging
from pathlib import Path
from pydantic import BaseModel, Field
from typing import List, Optional
import uuid
from datetime import datetime, timedelta, timedelta
#НОВОЕ
from fastapi.responses import StreamingResponse
from pymongo import MongoClient # DONOT REMOVE
import pandas as pd
import io
from fastapi import UploadFile, File
import openpyxl
from io import BytesIO


ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

# MongoDB connection
mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

# Create the main app without a prefix
app = FastAPI()

# Create a router with the /api prefix
api_router = APIRouter(prefix="/api")


# Define Models
class BorrowedBook(BaseModel):
    book_id: str
    book_title: str
    borrowed_date: datetime
    due_date: Optional[datetime] = None

class Student(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    first_name: str
    last_name: str
    class_name: str
    borrowed_books: List[BorrowedBook] = []
    created_at: datetime = Field(default_factory=datetime.utcnow)

class StudentCreate(BaseModel):
    first_name: str
    last_name: str
    class_name: str

class StudentUpdate(BaseModel):
    first_name: Optional[str] = None
    last_name: Optional[str] = None
    class_name: Optional[str] = None

class Book(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    title: str
    author: str
    quantity: int = 1
    borrowed_count: int = 0
    available: bool = True
    created_at: datetime = Field(default_factory=datetime.utcnow)

class BookCreate(BaseModel):
    title: str
    author: str
    quantity: int = 1

class BookUpdate(BaseModel):
    title: Optional[str] = None
    author: Optional[str] = None
    quantity: Optional[int] = None

class BorrowRequest(BaseModel):
    student_id: str
    book_id: str
    due_days: Optional[int] = 14  # Default 2 weeks

class ReturnRequest(BaseModel):
    student_id: str
    book_id: str

# Basic route
@api_router.get("/")
async def root():
    return {"message": "Librarian Assistant API"}

# Student endpoints
@api_router.post("/students", response_model=Student)
async def create_student(student: StudentCreate):
    student_dict = student.dict()
    student_obj = Student(**student_dict)
    await db.students.insert_one(student_obj.dict())
    return student_obj

@api_router.get("/students", response_model=List[Student])
async def get_all_students():
    students = await db.students.find().to_list(1000)
    return [Student(**student) for student in students]

@api_router.get("/students/class/{class_name}", response_model=List[Student])
async def get_students_by_class(class_name: str):
    students = await db.students.find({"class_name": class_name}).to_list(1000)
    return [Student(**student) for student in students]

@api_router.get("/classes")
async def get_all_classes():
    classes = await db.students.distinct("class_name")
    classes.sort()  
    return {"classes": classes}

@api_router.put("/students/{student_id}", response_model=Student)
async def update_student(student_id: str, student_update: StudentUpdate):
    update_data = {k: v for k, v in student_update.dict().items() if v is not None}
    if not update_data:
        raise HTTPException(status_code=400, detail="No valid fields provided for update")
    
    result = await db.students.update_one(
        {"id": student_id}, 
        {"$set": update_data}
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Student not found")
    
    updated_student = await db.students.find_one({"id": student_id})
    return Student(**updated_student)

@api_router.delete("/students/{student_id}")
async def delete_student(student_id: str):
    result = await db.students.delete_one({"id": student_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Student not found")
    return {"message": "Student deleted successfully"}

# Book endpoints
@api_router.post("/books", response_model=Book)
async def create_book(book: BookCreate):
    book_dict = book.dict()
    book_obj = Book(**book_dict)
    await db.books.insert_one(book_obj.dict())
    return book_obj

@api_router.get("/books", response_model=List[Book])
async def get_all_books():
    books = await db.books.find().to_list(1000)
    return [Book(**book) for book in books]

@api_router.put("/books/{book_id}", response_model=Book)
async def update_book(book_id: str, book_update: BookUpdate):
    update_data = {k: v for k, v in book_update.dict().items() if v is not None}
    if not update_data:
        raise HTTPException(status_code=400, detail="No valid fields provided for update")
    
    result = await db.books.update_one(
        {"id": book_id}, 
        {"$set": update_data}
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Book not found")
    
    updated_book = await db.books.find_one({"id": book_id})
    return Book(**updated_book)

@api_router.delete("/books/{book_id}")
async def delete_book(book_id: str):
    result = await db.books.delete_one({"id": book_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Book not found")
    return {"message": "Book deleted successfully"}

# Borrowing and returning endpoints
@api_router.post("/borrow")
async def borrow_book(borrow_request: BorrowRequest):
    # Check if book exists and is available
    book = await db.books.find_one({"id": borrow_request.book_id})
    if not book:
        raise HTTPException(status_code=404, detail="Book not found")
    
    book_obj = Book(**book)
    if book_obj.borrowed_count >= book_obj.quantity:
        raise HTTPException(status_code=400, detail="No copies available")
    
    # Check if student exists
    student = await db.students.find_one({"id": borrow_request.student_id})
    if not student:
        raise HTTPException(status_code=404, detail="Student not found")
    
    # Check if student already has this book
    student_obj = Student(**student)
    for borrowed_book in student_obj.borrowed_books:
        if borrowed_book.book_id == borrow_request.book_id:
            raise HTTPException(status_code=400, detail="Student already has this book")
    
    # Calculate due date
    due_date = datetime.utcnow() + timedelta(days=borrow_request.due_days)
    
    # Add book to student's borrowed books
    borrowed_book = BorrowedBook(
        book_id=borrow_request.book_id,
        book_title=book_obj.title,
        borrowed_date=datetime.utcnow(),
        due_date=due_date
    )
    
    await db.students.update_one(
        {"id": borrow_request.student_id},
        {"$push": {"borrowed_books": borrowed_book.dict()}}
    )
    
    # Increment borrowed count for book
    await db.books.update_one(
        {"id": borrow_request.book_id},
        {"$inc": {"borrowed_count": 1}}
    )
    
    return {"message": f"Book '{book_obj.title}' borrowed successfully", "due_date": due_date}

@api_router.post("/return")
async def return_book(return_request: ReturnRequest):
    # Check if student exists
    student = await db.students.find_one({"id": return_request.student_id})
    if not student:
        raise HTTPException(status_code=404, detail="Student not found")
    
    # Check if student has this book
    student_obj = Student(**student)
    book_found = False
    for borrowed_book in student_obj.borrowed_books:
        if borrowed_book.book_id == return_request.book_id:
            book_found = True
            book_title = borrowed_book.book_title
            break
    
    if not book_found:
        raise HTTPException(status_code=400, detail="Student doesn't have this book")
    
    # Remove book from student's borrowed books
    await db.students.update_one(
        {"id": return_request.student_id},
        {"$pull": {"borrowed_books": {"book_id": return_request.book_id}}}
    )
    
    # Decrement borrowed count for book
    await db.books.update_one(
        {"id": return_request.book_id},
        {"$inc": {"borrowed_count": -1}}
    )
    
    return {"message": f"Book '{book_title}' returned successfully"}

# Class management
class ClassCreate(BaseModel):
    name: str

# @api_router.post("/classes")
# async def create_class(class_data: ClassCreate):
#     # For now, we'll just return success since classes are created when students are added
#     # But we could store empty classes in the future if needed
#     return {"message": f"Class {class_data.name} is ready to accept students", "class_name": class_data.name}
@api_router.post("/classes")
async def create_class(class_data: ClassCreate):
    # Проверяем, есть ли уже класс с таким именем
    existing_class = await db.classes.find_one({"name": class_data.name})
    if existing_class:
        return {"message": f"Класс '{class_data.name}' уже существует"}

    # Создаём пустой класс
    new_class = {
        "name": class_data.name,
        "students": []  # пустой список учеников
    }

    await db.classes.insert_one(new_class)

    return {"message": f"Класс {class_data.name} создан", "class_name": class_data.name}


@api_router.get("/stats")
async def get_stats():
    # Get total students
    total_students = await db.students.count_documents({})
    
    # Get total books (sum of all quantities)
    books_cursor = db.books.find({})
    books = await books_cursor.to_list(1000)
    total_book_copies = sum(book.get('quantity', 1) for book in books)
    total_book_titles = len(books)
    
    # Get borrowed books count
    total_borrowed = sum(book.get('borrowed_count', 0) for book in books)
    available_copies = total_book_copies - total_borrowed
    
    # Get class counts
    pipeline = [
        {"$group": {"_id": "$class_name", "count": {"$sum": 1}}},
        {"$sort": {"_id": 1}}
    ]
    class_counts = await db.students.aggregate(pipeline).to_list(1000)
    
    return {
        "total_students": total_students,
        "total_book_titles": total_book_titles,
        "total_book_copies": total_book_copies,
        "available_copies": available_copies,
        "borrowed_copies": total_borrowed,
        "total_classes": len(class_counts),
        "class_counts": {item["_id"]: item["count"] for item in class_counts}
    }

# Include the router in the main app


app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()

# НОВОЕ
@api_router.get("/export_students")
async def export_students():
    cursor = db.students.find({}, {"_id": 0})
    students = await cursor.to_list(length=None)  

    if not students:
        return {"message": "Нет данных для экспорта"}

    # Превращаем данные в DataFrame
    df = pd.DataFrame(students)

    # Сохраняем Excel в память
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Ученики")
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=students.xlsx"}
    )

@api_router.get("/export_books")
async def export_books():
    cursor = db.books.find({}, {"_id": 0})
    books = await cursor.to_list(length=None)

    if not books:
        return {"message": "Нет данных для экспорта"}

    df = pd.DataFrame(books)
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Книги")
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=books.xlsx"}
    )


#NEW
@api_router.post("/students/import_excel")
async def import_students_excel(file: UploadFile = File(...)):
    if not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="Invalid file format. Please upload an Excel file.")

    try:
        contents = await file.read()
        workbook = openpyxl.load_workbook(BytesIO(contents), data_only=True)  # FIX
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Error reading Excel file: {str(e)}")

    sheet = workbook.active
    added_students = []
    skipped_students = []

    for row in sheet.iter_rows(min_row=2, values_only=True):
        last_name, first_name, class_name = row
        if not (last_name and first_name and class_name):
            continue  

        # Создаём класс, если его нет в коллекции "classes"
        existing_class = await db.classes.find_one({"name": class_name})
        if not existing_class:
            await db.classes.insert_one({"name": class_name})

        # Проверяем, есть ли ученик
        existing_student = await db.students.find_one({
            "first_name": first_name,
            "last_name": last_name,
            "class_name": class_name
        })
        if existing_student:
            skipped_students.append(f"{last_name} {first_name} ({class_name})")
            continue

        student_obj = {
            "id": str(uuid.uuid4()),
            "first_name": first_name,
            "last_name": last_name,
            "class_name": class_name
        }
        await db.students.insert_one(student_obj)
        added_students.append(f"{last_name} {first_name} ({class_name})")

    return {
        "added": added_students,
        "skipped": skipped_students
    }

@api_router.post("/books/import_excel")
async def import_books_excel(file: UploadFile = File(...)):
    if not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="Invalid file format. Please upload an Excel file.")

    try:
        contents = await file.read()
        workbook = openpyxl.load_workbook(BytesIO(contents), data_only=True)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Error reading Excel file: {str(e)}")

    sheet = workbook.active
    added_books = []
    updated_books = []

    for row in sheet.iter_rows(min_row=2, values_only=True):
        title, author, quantity = row
        if not (title and author and quantity):
            continue

        existing_book = await db.books.find_one({
            "title": title,
            "author": author
        })

        if existing_book:
            # Увеличиваем количество
            new_quantity = existing_book.get("quantity", 0) + int(quantity)
            await db.books.update_one(
                {"id": existing_book["id"]},
                {"$set": {"quantity": new_quantity}}
            )
            updated_books.append(f"{title} — {author} (+{quantity})")
        else:
            # Добавляем новую книгу
            book_obj = {
                "id": str(uuid.uuid4()),
                "title": title,
                "author": author,
                "quantity": int(quantity)
            }
            await db.books.insert_one(book_obj)
            added_books.append(f"{title} — {author} ({quantity})")

    return {
        "added": added_books,
        "updated": updated_books
    }


app.include_router(api_router)